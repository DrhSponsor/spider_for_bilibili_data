import os
import random
import time
import bilibili, fenbi, tieba, shiyebian
import openpyxl
import datetime
import requests

# Bark 推送 URL，请替换为你的 Bark API URL
BARK_URL = "https://api.day.app/gzThjYHuGrwDMFUnBsLM4b/"

def is_false(count):
    # 判断是否为字符串（即访问失败情况），返回 False 表示访问失败
    if isinstance(count, str):
        return "获取失败"
    else:
        return count

if __name__ == '__main__' :
    # 获取外部数据
    fenbi_count = is_false(fenbi.get_fenbi())  # 粉笔报考人数
    tieba_count = is_false(tieba.get_tieba())  # 贴吧帖子数量
    shiyebian_count = is_false(shiyebian.get_shiyebian())  # 事业编报名信息条数

    # 获取 BV 号列表
    bv_list = bilibili.getBV()

    # 存储所有视频的数据和失败的 BV 列表
    video_data_list = []
    failed_bv_list = []
    remove_count = 0

    # 遍历 BV 号列表，获取视频数据
    for bv in bv_list:
        print(f"正在获取 BV 号 {bv} 的数据...")
        try:
            # 获取视频数据字典
            video_data = bilibili.get_all_by_bv(bv)
            if video_data == '创作团队':
                bv_list.remove(bv)
                remove_count += 1
                continue
            video_row = [
                bv,
                video_data["view"],
                video_data["like"],
                video_data["share"],
                video_data["coin"],
                video_data["reply"],
                video_data["danmaku"],
                video_data["favorite"],
                video_data["fans"]
            ]
            print(video_row)
            # 将每个视频的数据加入列表
            video_data_list.append(video_row)
        except Exception as e:
            print(f"获取 {bv} 数据失败：", e)
            # 标记获取失败
            failed_bv_list.append(bv)
            video_row = [bv, "获取失败", "", "", "", "", "", "", ""]
            video_data_list.append(video_row)
        time.sleep(random.uniform(1, 3))

    if remove_count > 0:
        # 保存 BV 号列表到文件
        with open('./BV.txt', "w", encoding="utf-8") as file:
            file.write(str(bv_list))

    # 获取昨天的日期
    yesterday_date = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y%m%d")

    # 原始数据文件名
    raw_filename = f"{yesterday_date}哔哩哔哩考公视频原始数据.xlsx"
    summary_filename = "汇总.xlsx"

    # 创建原始数据 Excel 表格
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "原始数据"

    # 表头
    headers = [
        "BV号", "播放量", "点赞数", "分享量", "投币量", "评论数", "弹幕数", "收藏量", "UP主粉丝量"
    ]
    ws.append(headers)

    # 写入每个视频的数据
    for row in video_data_list:
        ws.append(row)

    # 保存原始数据表格
    wb.save(raw_filename)
    print(f"原始数据保存为：{raw_filename}")

    # 创建汇总数据 Excel 表格
    try:
        # 如果汇总表格存在，加载它；否则创建新的
        if os.path.exists(summary_filename):
            summary_wb = openpyxl.load_workbook(summary_filename)
            summary_ws = summary_wb.active
        else:
            summary_wb = openpyxl.Workbook()
            summary_ws = summary_wb.active
            summary_ws.title = "汇总数据"
            # 汇总表头
            summary_headers = [
                "日期", "总播放量", "总点赞数", "总分享量", "总投币量", "总评论数", "总弹幕数", "总收藏量",
                "粉笔报考人数", "贴吧帖子数", "事业编报名条数",
                "粉笔增量", "贴吧增量", "事业编增量"
            ]
            summary_ws.append(summary_headers)

        # 计算汇总数据
        total_view = sum(row[1] if isinstance(row[1], int) else 0 for row in video_data_list)
        total_like = sum(row[2] if isinstance(row[2], int) else 0 for row in video_data_list)
        total_share = sum(row[3] if isinstance(row[3], int) else 0 for row in video_data_list)
        total_coin = sum(row[4] if isinstance(row[4], int) else 0 for row in video_data_list)
        total_reply = sum(row[5] if isinstance(row[5], int) else 0 for row in video_data_list)
        total_danmaku = sum(row[6] if isinstance(row[6], int) else 0 for row in video_data_list)
        total_favorite = sum(row[7] if isinstance(row[7], int) else 0 for row in video_data_list)

        # 获取上一天的数据（如果有）
        if summary_ws.max_row > 1:
            last_row = summary_ws[summary_ws.max_row]
            last_fenbi = last_row[8].value if isinstance(last_row[8].value, int) else 0
            last_tieba = last_row[9].value if isinstance(last_row[9].value, int) else 0
            last_shiyebian = last_row[10].value if isinstance(last_row[10].value, int) else 0
        else:
            last_fenbi, last_tieba, last_shiyebian = 0, 0, 0

        # 计算增量
        fenbi_increase = fenbi_count if last_fenbi == 0 else fenbi_count - last_fenbi
        tieba_increase = tieba_count if last_tieba == 0 else tieba_count - last_tieba
        shiyebian_increase = shiyebian_count if last_shiyebian == 0 else shiyebian_count - last_shiyebian

        # 添加汇总数据行
        summary_row = [
            yesterday_date, total_view, total_like, total_share, total_coin,
            total_reply, total_danmaku, total_favorite,
            fenbi_count, tieba_count, shiyebian_count,
            fenbi_increase, tieba_increase, shiyebian_increase
        ]
        summary_ws.append(summary_row)

        # 保存汇总表格
        summary_wb.save(summary_filename)
        print(f"汇总数据保存为：{summary_filename}")

        # 发送 Bark 推送
        data = {'body':f"失败视频：{len(failed_bv_list)},粉笔：{fenbi_count},贴吧：{tieba_count},事业编：{shiyebian_count}"}
        requests.post(BARK_URL,data=data)

    except Exception as e:
        print("汇总表格更新失败：", e)
